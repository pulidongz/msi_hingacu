from openpyxl.utils.exceptions import InvalidFileException
from abc import ABCMeta, abstractmethod
from etl.models import Workbook, DataExtractionConfiguration


class ExtractionProcess:
    # Generalized Process for Validating and Extracting Data from a worksheet
    __metaclass__ = ABCMeta

    def __init__(self, config, workbook):
        # assumption here is that the wb has already been checked to contain the config's requirements
        self.workbook = workbook
        self.config = config
        print(self.config)

    def validate_config(self):
        #override me
        return

    def get_worksheet(self):
        # Initialize Worksheet
        wb = self.workbook.get_workbook()
        sheet_name = self.config.sheet_name
        try:
            ws = wb[sheet_name]
        except KeyError:
            message = 'Sheet "' + sheet_name + '" does not exist.'
            raise Workbook.DoesNotExist(message)
        return ws

    @abstractmethod
    def extract(self):
        ''' To override '''
        #ws = self.get_worksheet()
        #return data
        pass

    def validate(self, data):
        # validate extracted data with given form_class

        rules = self.config.rules['fields']
        form_class = self.config.get_validation_form()

        form = form_class(data, rules=rules, file=self.workbook)
        result = {
            'data': None,
            'errors': None
        }
        if form.is_valid():
            result['data'] = form.cleaned_data
        else:
            result['errors'] = form.errors.as_data()
        return result

    def transform(self, data):
        # transform data in preparation for saving to database
        #print(data)
        return data

    def load(self, data):
        # save data into database
        print(data)

    def run(self):
        #extract raw data based on config
        raw_data = self.extract()

        results = self.validate(raw_data)

        #save results found into database
        data = self.transform(results)
        self.load(data)


class FormDataExtraction(ExtractionProcess):
    #Extract Values from a worksheet formatted as a form
    #config will have a cell field for every field_name

    def extract(self):
        ws = self.get_worksheet()

        data = {}
        for field in self.config.rules['fields']:
            value = ws[field['cell']].value
            data[field['field_name']] = value

        print("RAW", data)
        return data


class RowDataExtraction(ExtractionProcess):
    #Extract values from a worksheet with a table format
    #config will have a column index for each field_name

    def convert_entry_to_values(self, entry):
        entry_values = []
        for cell in entry:
            if cell.hyperlink: # use hyperlink value instead of text value if available
                value = cell.hyperlink.target
            else:
                value = cell.value
            # cast integers as integers. openpyxl reads all numbers as floats.
            if type(value) == float:
                if value.is_integer():
                    value = int(value)
            entry_values.append(value)
        return entry_values

    def convert_entry_to_dict(self, entry):
        fields = self.config.rules['fields']
        entry_values = self.convert_entry_to_values(entry)
        data = {}
        for field in fields:
            data[field['field_name']] = entry_values[field['column_index']]
        print(data)
        return data

    def extract(self):
        ws = self.get_worksheet()
        min_row = self.config.rules['min_row']
        max_row = self.config.rules['max_row']
        min_col = self.config.rules['min_col']
        max_col = self.config.rules['max_col']

        #TO DO VALIDATE RULES WITH NUMBER OF COLUMNS AND COLUMN NAMES
        #BEFORE EXTRACTING

        table_data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            if any(cell.value for cell in row): #process only non-blank rows
                data = self.convert_entry_to_dict(row)
                table_data.append(data)

        return table_data

    def run(self):
        #extract raw data based on config
        raw_table_data = self.extract()

        for entry in raw_table_data: #raw data is a list, validate per element
            cleaned_data = self.validate(entry)
            data = self.transform(cleaned_data)
            self.load(data)


class ColumnDataExtraction(RowDataExtraction):
    #Extract values from a worksheet with a table format where each entry in a column
    #config will have a row_index for each field_name

    def extract(self):
        ws = self.get_worksheet()
        min_row = self.config.rules['min_row']
        max_row = self.config.rules['max_row']
        min_col = self.config.rules['min_col']
        max_col = self.config.rules['max_col']

        #TO DO VALIDATE RULES WITH NUMBER OF COLUMNS AND COLUMN NAMES
        #BEFORE EXTRACTING

        table_data = []
        for col in ws.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            if any(cell.value for cell in col): #process only non-blank columns
                data = self.convert_entry_to_dict(col)
                table_data.append(data)

        return table_data
