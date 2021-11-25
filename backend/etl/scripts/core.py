import json
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Fill
from openpyxl.writer.excel import save_virtual_workbook
from etl.models import ETLFile, ETLFileRow, ExtractedData
from etl.forms import DCRForm
from etl.utils import get_printable


class ETLProcess:
    # Generalized Object for ETLProcess

    def __init__(self, dcp, form_class, result_sheet):
        self.dcp = dcp
        self.sheet_name = dcp.sheet_name
        self.key_mapping = dcp.requirements
        self.form_class = form_class
        self.result_sheet = result_sheet

    def extract(self, etlfile):
        try:
            # extract per row data from sheet based on given data mapping
            raw_data = etlfile.get_data(
                self.key_mapping,
                sheet_name=self.sheet_name)
        except ETLFile.DoesNotExist as file_error:
            etlfile.add_file_error(file_error, self.sheet_name)
            etlfile.status = ETLFile.STATUS_FAILED
            etlfile.save()
            return #stop here

        # debug line
        # print(self.sheet_name, raw_data)

        # delete all ETLFileRows associated to ETLFile and Sheet name
        ETLFileRow.objects.filter(etlfile=etlfile, sheet_name=self.sheet_name).delete()

        # assign a number to each row
        for row in raw_data:
            row_index = row['ROW_INDEX']
            #print("ROW", row_index, "PROCESSED")
            #VALIDATION
            form = self.form_class(row, dcr=self.key_mapping, etlfile=etlfile)
            if form.is_valid():
                self.transform_and_load(etlfile, form, row_index)
            else:
                # mark row as invalid if errors found
                form_errors = form.errors.as_data()
                etlfile.add_errors(form_errors, row_index, self.dcp, self.result_sheet)

        # save errors and update status based on errors
        if etlfile.errors:
            etlfile.status = ETLFile.STATUS_FAILED
            etlfile.save()

    def transform_and_load(self, etlfile, form, row_index):
        # OVERRIDE THIS METHOD FOR CUSTOM TRANSFORMATION PER DCP
        form.save(self.sheet_name, row_index)

        # highlight warnings in result sheet
        for field_name in form.warnings:
            etlfile.highlight_warning(row_index, field_name, self.dcp, self.result_sheet)
        #print('Transform and Load', etlfile, row_index, json_data)

class FormSheetETLProcess:
    # Generalized Process for Validating and Extracting Data from a datasheet

    def __init__(self, etlfile, dcp):
        self.etlfile = etlfile
        self.dcp = dcp
        self.requirements = dcp.get_dcp_configuration()
        self.form_class = dcp.get_validation_form()
        #self.result_sheet = result_sheet

    def extract(self):
        # Initialize Worksheet
        try:
            wb = self.etlfile.get_workbook()
            print('SHEET NAMES', wb.sheetnames)
            sheet_name = self.dcp.sheet_name
            try:
                ws = wb[sheet_name]
            except KeyError:
                message = 'Sheet "' + sheet_name + '" does not exist.'
                raise ETLFile.DoesNotExist(message)
        except InvalidFileException as e:
            raise e

        # Get worksheet value per field in DCP Configuration
        data = {}
        for field in self.requirements:
            cell_coordinate = field["field_cell"].split(":")[0] # only top left reference has value for merged cells
            cell = ws[cell_coordinate]
            print(cell)
            data[field["field_name"]] = cell.value

        return data

    def validate(self, data):
        # validate extracted data with given form_class

        dcr = {} # convert list to dict for now. TO DO: update DCRform to accept a list
        for requirement in self.requirements:
            dcr[requirement['field_name']] = requirement

        form = self.form_class(data, dcr=dcr, etlfile=self.etlfile)
        if form.is_valid():
            return form.cleaned_data
        else:
            return False

    def transform(self, data):
        # transform data in preparation for saving to database
        print(data)
        return data

    def load(self, data):
        # save data into database
        saved_data = ExtractedData.objects.create(
            etlfile = self.etlfile,
            sheet_name = self.dcp.sheet_name,
            data = data
        )

    def run(self):
        raw_data = self.extract()
        data = self.validate(raw_data)
        data = self.transform(data)
        self.load(data)


def process_dcp(dcp, etlfile, result_sheet):
    mapping = dcp.requirements
    sheet_name = dcp.sheet_name
    form_class = DCRForm

    if settings.DEBUG:
        print("PROCESS DCP: ", dcp)

    if dcp.requirements['configuration_type'] == 'form':
        etlprocess = FormSheetETLProcess(etlfile, dcp)
        etlprocess.run()
    else:
        etlprocess = ETLProcess(dcp, form_class, result_sheet)
        etlprocess.extract(etlfile)


def process_etlfile(etlfile):
    # delete all proxy data related to file
    etlfile.errors = []
    etlfile.status = ETLFile.STATUS_PROCESSING
    etlfile.save()

    dcps = etlfile.dcp_collection.datacapturepoint_set.all().order_by('date_created')
    #print("DCPS", dcps)

    #initialize result workbook
    result_wb = etlfile.get_workbook() #load a copy of the wb

    for dcp in dcps:
        try:
            result_sheet = result_wb[dcp.sheet_name]
            result_sheet.error_column = result_sheet.max_column + 2
            error_column_header = result_sheet.cell(row=1, column=result_sheet.error_column)
            error_column_header.value = "ERRORS FOUND"
            process_dcp(dcp, etlfile, result_sheet)
        except Exception as e:
            etlfile.add_file_error(e, dcp.sheet_name)
            etlfile.status = ETLFile.STATUS_FAILED
            etlfile.save()
            raise e

    # save result workbook
    file_name = "%s_%s_results.xlsx" % (etlfile.pk, etlfile.dcp_collection.code)
    etlfile.result_file.save(file_name, ContentFile(save_virtual_workbook(result_wb)))

    if etlfile.errors:
        etlfile.status = ETLFile.STATUS_FAILED
    else:
        etlfile.status = ETLFile.STATUS_COMPLETED
    etlfile.save()
