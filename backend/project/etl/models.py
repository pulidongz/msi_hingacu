import os
from django.db import models
from django.utils.translation import gettext, gettext_lazy as _
from django.utils.text import slugify
from django.utils.module_loading import import_string
from django.core.files.base import ContentFile
from django.urls import reverse
from django.contrib.auth.models import User
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django_extensions.db.fields import AutoSlugField
from django_q.models import Task
from django_q.tasks import async_task, result, fetch
from openpyxl import load_workbook
from openpyxl.styles import Fill, PatternFill
from openpyxl.comments import Comment
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from etl.utils import convert_row_to_dict, get_printable
from etl.validators import validate_workbook_extension, validate_workbook_attachment


class TimeStampModel(models.Model):
    date_created = models.DateTimeField(auto_now_add=True)
    date_updated = models.DateTimeField(auto_now=True)

    class Meta:
        abstract = True


class DCPCollection(TimeStampModel):
    """AKA the Business Process Model"""

    code = models.CharField(primary_key=True, max_length=100)
    name = models.CharField(max_length=100, null=True, blank=True)
    details = models.TextField(null=True, blank=True)

    class Meta:
        verbose_name = "DCP Collection"
        verbose_name_plural = "DCP Collections (Business Process Models)"

    def __str__(self):
        if self.name:
            return self.name
        return self.code


class DataCapturePoint(TimeStampModel):
    collection = models.ForeignKey(DCPCollection, on_delete=models.CASCADE)
    code = models.CharField(max_length=100)
    label = models.CharField(max_length=100)
    sheet_name = models.CharField(max_length=50)
    requirements = models.JSONField()

    class Meta:
        verbose_name = "DCP"
        verbose_name_plural = "DCP Configurations"
        unique_together = [['code', 'collection']]
        indexes = [
            models.Index(fields=['code', 'collection']),
        ]

    def __str__(self):
        return self.label

    def get_column_name(self, field_name):
        field = self.requirements[field_name]
        return field['column_name']

    def requirements_list(self):
        # return DCP requirements in list form
        l = []
        sorted_keys = sorted(self.requirements.keys(), key=lambda x:x.lower())
        for key in sorted_keys:
            requirement = self.requirements[key]
            requirement['field_name'] = key
            l.append(requirement)
        return l

    def get_dcp_configuration(self):
        if self.requirements["configuration_type"] == "form":
            return self.requirements["fields"]
        else:
            return self.requirements_list()

    def get_dcp_type(self):
        return self.requirements["configuration_type"]

    def is_dcp_form_type(self):
        return self.requirements["configuration_type"] == "form"

    def get_validation_form(self):
        form = import_string('etl.forms.DCRForm')
        return form


class ETLFile(TimeStampModel):
    STATUS_QUEUED = 0
    STATUS_FAILED = 1
    STATUS_COMPLETED = 2
    STATUS_PROCESSING = 3
    STATUS_CHOICES = (
        (STATUS_QUEUED, 'Queued'),
        (STATUS_FAILED, 'With Errors'),
        (STATUS_COMPLETED, 'Completed'),
        (STATUS_PROCESSING, 'Processing')
    )

    dcp_collection = models.ForeignKey(DCPCollection, on_delete=models.CASCADE,
        verbose_name="Project/Component")
    uploader = models.ForeignKey(User, on_delete=models.CASCADE)
    file = models.FileField(upload_to='etl')
    result_file = models.FileField(upload_to='results', null=True)
    task_id = models.CharField(max_length=32, null=True)
    errors = models.JSONField(null=True)
    status = models.IntegerField(default=0, choices=STATUS_CHOICES)

    class Meta:
        verbose_name_plural = "ETL Files"
        verbose_name = "ETL File"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.column_names = []
        self.workbook = None
        self.column_map = {}

    def __str__(self):
        return str(self.dcp_collection.name) + ": " + str(self.pk)

    def get_filename(self):
        return os.path.basename(self.file.name)

    def get_data(self, key_mapping, sheet_name=None):
        return self.load_datasheet(key_mapping, sheet_name)

    def get_workbook(self):
        # so we only load the workbook once in memory
        if self.workbook is None: #first access
            self.workbook = load_workbook(filename=self.file.path, data_only=True)
        return self.workbook

    def get_column_map(self, sheet_name):
        if sheet_name not in self.column_map:
            self.column_map[sheet_name] = {}
            wb = self.get_workbook()
            column_row = wb[sheet_name][1]
            for cell in column_row:
                if cell.value:
                    column_name = get_printable(cell.value)
                    column_key = column_name.lower() #lower case keys
                    self.column_map[sheet_name][column_key] = cell.column_letter
        return self.column_map[sheet_name]

    def get_column_coordinate(self, dcp, field_name):
        column_map = self.get_column_map(dcp.sheet_name)
        column_name = dcp.get_column_name(field_name)
        column_key = get_printable(column_name.lower())
        #print('COLUMN KEY', column_key)
        #print('COLUMN MAP', column_map)
        return column_map[column_key]

    def load_datasheet(self, key_mapping, sheet_name=None):
        try:
            wb = self.get_workbook()
            print('SHEET NAMES', wb.sheetnames)
            if sheet_name:
                try:
                    ws = wb[sheet_name]
                except KeyError:
                    message = 'Sheet "' + sheet_name + '" does not exist.'
                    raise self.DoesNotExist(message)
            else:
                ws = wb[wb.active]

            self.data = []
            row_index = 0
            for row in ws.iter_rows():
                row = self.convert_row_to_values(row)
                if row_index == 0: #first row with column names
                    self.column_names = list(filter(None, row))
                    first = False
                    if settings.DEBUG:
                        print("COLUMNS", self.column_names)
                    #fail in advance if columns are incomplete
                    self.check_if_columns_are_complete(key_mapping)
                else:
                    if not row.count(None) == len(row): #Skip empty rows
                        data = self.convert_row_to_dict(row, key_mapping)
                        data['ROW_INDEX'] = row_index + 1
                        self.data.append(data)
                row_index = row_index + 1
            return self.data
        except InvalidFileException as e:
            raise e

    def extract_data(self, dcp):
        try:
            wb = self.get_workbook()
            print('SHEET NAMES', wb.sheetnames)
            if sheet_name:
                try:
                    ws = wb[sheet_name]
                except KeyError:
                    message = 'Sheet "' + sheet_name + '" does not exist.'
                    raise self.DoesNotExist(message)
            else:
                ws = wb[wb.active]

            self.data = []
            row_index = 0
            for row in ws.iter_rows():
                row = self.convert_row_to_values(row)
                if row_index == 0: #first row with column names
                    self.column_names = list(filter(None, row))
                    first = False
                    if settings.DEBUG:
                        print("COLUMNS", self.column_names)
                    #fail in advance if columns are incomplete
                    self.check_if_columns_are_complete(key_mapping)
                else:
                    if not row.count(None) == len(row): #Skip empty rows
                        data = self.convert_row_to_dict(row, key_mapping)
                        data['ROW_INDEX'] = row_index + 1
                        self.data.append(data)
                row_index = row_index + 1
            return self.data
        except InvalidFileException as e:
            raise e

    def convert_row_to_values(self, row):
        cell_values = []
        for cell in row:
            if cell.hyperlink: # use hyperlink value instead of text value if available
                value = cell.hyperlink.target
            else:
                value = cell.value
            # cast integers as integers. openpyxl reads all numbers as floats.
            if type(value) == float:
                if value.is_integer():
                    value = int(value)
            cell_values.append(value)
        #print(cell_values)
        return cell_values

    def convert_row_to_dict(self, row, key_mapping):
        column_mapped_data = convert_row_to_dict(row, self.column_names)
        data = {}
        for key in key_mapping.keys():
            try:
                data[key] = column_mapped_data[key_mapping[key]['column_name'].strip().lower()]
            except KeyError:
                message = 'Column "' + key_mapping[key]['column_name'] + '" does not exist.'
                raise self.DoesNotExist(message)
        #if settings.DEBUG:
        #    print("COLUMN_LABELS", column_mapped_data)
        #    print("FIELD_LABELS", data)
        return data

    def check_if_columns_are_complete(self, key_mapping):
        missing_columns = False
        message = "The following columns are missing:"
        column_names_lower = [x.strip().lower() for x in self.column_names]
        for key in key_mapping.keys():
            #print(key, key_mapping)
            if not isinstance(key_mapping[key], dict):
                raise Exception('Faulty Configuration in DCR.')
            if not key_mapping[key]['column_name'].strip().lower() in column_names_lower:
                message = message + ' "' + key_mapping[key]['column_name'] + '",'
                missing_columns = True
        if missing_columns:
            message = message.strip(',') + '.'
            raise self.DoesNotExist(message)

    def add_errors(self, form_errors, row_index, dcp, result_sheet):
        # FOR FORM VALIDATION ERRORS

        # highlight entire row in result sheet
        self.highlight_row(row_index, result_sheet)

        # prep a variable for error message to be embedded in result sheet
        sheet_error_message = ""

        # store each error in etlfile error json
        for field in form_errors.keys():
            for error in form_errors[field]:
                error_details = {
                    'row': row_index,
                    'column': field,
                    'error': str(error.message),
                    'sheet': dcp.sheet_name,
                    'sheet_slug': slugify(dcp.sheet_name)
                }
                # add errors to json field
                self.errors.append(error_details)

                sheet_error_message = sheet_error_message + field + ": " + str(error.message) + " "

                # highlight cell errors in result sheet
                column_letter = self.get_column_coordinate(dcp, field)
                self.hightlight_cell_error(row_index, column_letter, result_sheet)

        # write errors to a new column of result sheet
        error_cell = result_sheet.cell(row=row_index, column=result_sheet.error_column)
        error_cell.value = sheet_error_message

    def highlight_row(self, row_number, result_sheet):
        rowFill = PatternFill(fgColor="FFCCCB", fill_type="solid")
        row = result_sheet[row_number]
        for cell in row:
            cell.fill = rowFill

    def hightlight_cell_error(self, row_number, column_letter, result_sheet):
        cell_coordinate = f"{column_letter}{row_number}"
        redFill = PatternFill(fgColor="FF0000", fill_type="solid")
        result_sheet[cell_coordinate].fill = redFill

    def highlight_warning(self, row_number, field_name, dcp, result_sheet):
        column_letter = self.get_column_coordinate(dcp, field_name)
        cell_coordinate = f"{column_letter}{row_number}"
        yellowFill = PatternFill(fgColor="FFFF00", fill_type="solid")
        result_sheet[cell_coordinate].fill = yellowFill

    def add_sheet_warnings(self, warnings, row_index, sheet_name):
        # for soft requirements
        for warning in warnings:
            warning_details = {
                'row': row_index,
                'column': field,
                'error': str(error.message),
                'sheet': sheet_name,
                'sheet_slug': slugify(sheet_name)
            }
            self.errors.append(error_details)


    def add_file_error(self, error, sheet_name):
        # FOR FILE ERRORS
        error_details = {
            'row': 'N/A',
            'column': 'N/A',
            'error': str(error),
            'sheet': sheet_name,
            'sheet_slug': slugify(sheet_name)
        }
        self.errors.append(error_details)

    def get_errors(self):
        sheet_map = {}
        for error in self.errors:
            if error['sheet_slug'] not in sheet_map:
                sheet_map[error['sheet_slug']] = []
            sheet_map[error['sheet_slug']].append(error)
        return sheet_map

    def get_report(self):
        sorted_sheets = {}
        dcps = self.dcp_collection.datacapturepoint_set.all().order_by('date_created')
        for dcp in dcps:
            sorted_sheets[dcp.sheet_name] = []
        for error in self.errors:
            if error['sheet'] not in sorted_sheets:
                sorted_sheets[error['sheet']] = []
            sorted_sheets[error['sheet']].append(error)
        sheet_list = []
        i = 0
        for key in sorted_sheets:
            dcp = {
                'dcp': key,
                'index': i,
                'error_list': sorted_sheets[key],
                'data': ETLFileRow.objects.filter(etlfile=self, sheet_name=key)
            }
            i = i + 1
            sheet_list.append(dcp)
        return sheet_list


class ETLFileRow(TimeStampModel):
    etlfile = models.ForeignKey(ETLFile, on_delete=models.CASCADE)
    sheet_name = models.CharField(max_length=64)
    sheet_slug = AutoSlugField(null=True, default=None, allow_duplicates=True, populate_from='sheet_name')
    number = models.IntegerField()
    data = models.JSONField()
    warnings = models.JSONField(null=True)

    class Meta:
        verbose_name = "ETL File Row Data"
        verbose_name_plural = "ETL File Row Data"

    def __str__(self):
        return str(self.etlfile.pk) + ' ' + self.sheet_name + ': ' + str(self.number)


class GoogleDriveFile(TimeStampModel):
    drive_id = models.TextField(null=True)
    drive_url = models.URLField()
    file = models.FileField(upload_to='drive', null=True)
    downloaded = models.BooleanField(default=False)

    def __str__(self):
        return self.drive_url

    def filename(self):
        return os.path.basename(self.file.name)


class KeyValueStore(TimeStampModel):
    etlfile = models.ForeignKey(ETLFile, on_delete=models.CASCADE)
    sheet_name = models.CharField(max_length=64)
    code = models.TextField()
    data = models.JSONField(null=True)

    def __str__(self):
        return f"{self.sheet_name}: {self.code}"


class WorkbookConfiguration(TimeStampModel):
    """Collection of Sheet Configurations"""

    component = models.CharField(max_length=200, null=True, blank=True)
    code = models.CharField(max_length=20)
    name = models.CharField(max_length=100, null=True, blank=True)
    version = models.IntegerField(default=0)
    description = models.TextField(null=True, blank=True)
    publisher_class = models.TextField(null=True, blank=True)

    class Meta:
        unique_together = [['code', 'version']]
        indexes = [
            models.Index(fields=['code', 'version']),
        ]
        ordering = ["name"]

    def __str__(self):
        if self.name:
            return f"{self.name}"
        return self.code

    def get_publisher_class(self):
        if self.publisher_class:
            return import_string(self.publisher_class)
        return None

    def get_sheet_configurations(self):
        return self.worksheetconfiguration_set.all().order_by('sheet_order')


class WorksheetConfiguration(TimeStampModel): #Worksheet Model
    workbook_configuration = models.ForeignKey(WorkbookConfiguration, on_delete=models.CASCADE)
    sheet_name = models.CharField(max_length=100)
    sheet_order = models.IntegerField(default=0)

    class Meta:
        ordering = ["workbook_configuration", "sheet_order"]

    def __str__(self):
        return f"{self.code()}: {self.sheet_name}"

    def code(self):
        return f"{self.workbook_configuration.code}-SHEET-{self.sheet_order}"

    def get_extraction_configurations(self):
        return self.dataextractionconfiguration_set.all().order_by('section_order')


class DataExtractionConfiguration(TimeStampModel):
    TYPE_FORM = 'form'
    TYPE_TABLE_ROWS = 'table_rows'
    TYPE_TABLE_COLUMNS = 'table_columns'
    EXTRACTION_CHOICES = (
        (TYPE_FORM, 'Form'),
        (TYPE_TABLE_ROWS, 'Table Rows'),
        (TYPE_TABLE_COLUMNS, 'Table Columns')
    )

    worksheet_configuration = models.ForeignKey(WorksheetConfiguration, on_delete=models.CASCADE)
    name = models.CharField(max_length=200)
    section_order = models.IntegerField(default=0)
    min_row = models.IntegerField(default=1)
    max_row = models.IntegerField(default=10)
    min_col = models.IntegerField(default=1)
    max_col = models.IntegerField(default=10)
    extraction_type = models.CharField(max_length=20, default=TYPE_FORM, choices=EXTRACTION_CHOICES)
    custom_validation_form = models.CharField(max_length=200, blank=True, null=True)
    custom_etl_class = models.CharField(max_length=200, blank=True, null=True)

    class Meta:
        verbose_name = "Worksheet Section"
        verbose_name_plural = "Worksheet Sections"

    def __str__(self):
        return f"{self.worksheet_configuration.code()}: {self.name}"

    def get_validation_form(self):
        #returns the validation form used for this extraction
        if self.custom_validation_form:
            return import_string(self.custom_validation_form)
        return import_string('etl.forms.DefaultValidationForm')

    def get_etl_class(self):
        #returns the process class used for this extraction configuration
        if self.custom_etl_class:
            return import_string(self.custom_etl_class)
        if self.extraction_type == DataExtractionConfiguration.TYPE_FORM:
            return import_string('etl.workbooks.extraction.FormDataExtraction')
        elif self.extraction_type == DataExtractionConfiguration.TYPE_TABLE_ROWS:
            return import_string('etl.workbooks.extraction.RowDataExtraction')
        elif self.extraction_type == DataExtractionConfiguration.TYPE_TABLE_COLUMNS:
            return import_string('etl.workbooks.extraction.ColumnDataExtraction')
        return import_string('etl.workbooks.extraction.ExtractionProcess')

    def get_field_rules(self):
        return self.datafield_set.all()

    def is_form_config(self):
        return self.extraction_type == self.TYPE_FORM

    def is_table_rows_config(self):
        return self.extraction_type == self.TYPE_TABLE_ROWS

    def is_table_columns_config(self):
        return self.extraction_type == self.TYPE_TABLE_COLUMNS


class DataField(TimeStampModel):
    TYPE_TEXT = "text"
    TYPE_INTEGER = "integer"
    TYPE_DECIMAL = "decimal"
    TYPE_DATE = "date"
    TYPE_TIME = "time"
    TYPE_BOOLEAN = "boolean"
    TYPE_COORDINATE = "coordinate"
    TYPE_CHOICES = (
        (TYPE_TEXT, 'Text'),
        (TYPE_INTEGER, 'Integer'),
        (TYPE_DECIMAL, 'Decimal'),
        (TYPE_DATE, 'Date'),
        (TYPE_TIME, 'Time'),
        (TYPE_BOOLEAN, 'Boolean'),
        (TYPE_COORDINATE, 'Coordinate'),
    )

    section = models.ForeignKey(DataExtractionConfiguration, on_delete=models.CASCADE)
    field_name = models.CharField(max_length=200)
    field_type = models.CharField(max_length=20, choices=TYPE_CHOICES, default=TYPE_TEXT)
    field_index = models.IntegerField(default=0)
    required = models.BooleanField(default=True)
    cell = models.CharField(max_length=10, null=True, blank=True)

    class Meta:
        ordering = ['field_index']


    def __str__(self):
        return f"{self.section.name}: {self.field_name} field."


class Workbook(TimeStampModel):
    STATUS_NEW = 'new'
    STATUS_QUEUED = 'queued'
    STATUS_FAILED = 'failed'
    STATUS_COMPLETED = 'completed'
    STATUS_CORRECTIONS_NEEDED = 'corrections_needed'
    STATUS_PROCESSING = 'processing'
    STATUS_CHOICES = (
        (STATUS_NEW, 'Newly Uploaded'),
        (STATUS_QUEUED, 'Queued'),
        (STATUS_FAILED, 'Failed'),
        (STATUS_COMPLETED, 'Completed'),
        (STATUS_CORRECTIONS_NEEDED, 'Needs Corrections'),
        (STATUS_PROCESSING, 'Processing')
    )

    configuration = models.ForeignKey(WorkbookConfiguration, on_delete=models.SET_NULL, null=True)
    uploader = models.ForeignKey(User, on_delete=models.SET_NULL, null=True)
    file = models.FileField(upload_to='workbooks', validators=[validate_workbook_extension])
    attachments = models.FileField(upload_to='attachments', validators=[validate_workbook_attachment], blank=True, null=True)
    file_with_corrections = models.FileField(upload_to='corrections', blank=True, null=True)
    status = models.CharField(max_length=30, default=STATUS_NEW, choices=STATUS_CHOICES)

    def __init__(self, *args, **kwargs):
        # initialize workbook internal variables
        super().__init__(*args, **kwargs)
        self.workbook = None
        self.result_wb = None

    def __str__(self):
        return f"{self.configuration.name} {self.pk}"

    def get_workbook(self):
        # so we only load the workbook once
        if self.workbook is None: #first access
            self.workbook = load_workbook(filename=self.file.path, data_only=True)
        return self.workbook

    def get_workbook_copy(self):
        # so we only load the result workbook once
        if self.result_wb is None: #first access
            self.result_wb = load_workbook(filename=self.file.path)
        return self.result_wb

    def process(self):
        #queue the workbook for async processing if not already queued
        if self.status not in (self.STATUS_QUEUED, self.STATUS_PROCESSING):
            async_task('etl.workbooks.process', self.pk, hook='etl.workbooks.complete')

    def get_filename(self):
        return os.path.basename(self.file.name)

    def get_filename_checked(self):
        return os.path.basename(self.file_with_corrections.name)

    def code(self):
        return f"{self.configuration.code}-{self.pk}"

    def get_available_sheets(self):
        workbook = self.get_workbook()
        return workbook.sheetnames

    def sheet_is_available(self, sheet_name):
        available_sheets = self.get_available_sheets()
        return sheet_name in available_sheets

    def is_ready(self):
        return self.status not in (self.STATUS_QUEUED, self.STATUS_PROCESSING, self.STATUS_NEW)

    def is_complete(self):
        return self.status == self.STATUS_COMPLETED

    def get_status_class(self):
        if self.status == self.STATUS_COMPLETED:
            return 'text-success'
        elif self.status in (self.STATUS_QUEUED, self.STATUS_PROCESSING, self.STATUS_NEW):
            return 'text-info'
        return 'text-danger'

    def get_validation_status(self):
        if self.status == self.STATUS_COMPLETED:
            return 'passed'
        elif self.status == self.STATUS_FAILED:
            return 'failed'
        return 'ongoing'

    def get_absolute_url(self):
        return reverse('etl:workbook-detail', args=[str(self.pk)])

    def has_errors(self):
        return self.extractionerror_set.exists()

    def has_field_errors(self):
        return self.extractionerror_set.filter(field_error=True).exists()

    def has_non_field_errors(self):
        return self.extractionerror_set.filter(field_error=False).exists()

    def non_field_errors(self):
        return self.extractionerror_set.filter(field_error=False)

    def add_error(self, configuration, field_name, error_message):
        #form validation related errors
        ExtractionError.objects.create(
            workbook = self.workbook,
            configuration = configuration,
            field_name = field_name,
            error = error_message
        )


    def add_file_error(self, error_message):
        # file errors are errors related to the config like missing sheets or fields
        ExtractionError.objects.create(
            workbook = self,
            field_error = False,
            error = error_message
        )

    def highlight_row(self, row_number, result_sheet):
        rowFill = PatternFill(fgColor="FFCCCB", fill_type="solid")
        row = result_sheet[row_number]
        for cell in row:
            cell.fill = rowFill

    def generate_checked_wb(self):
        # create a workbook copy where error fields are highlighted.
        checked_wb = self.get_workbook_copy()

        # remove comments and set all cells to transparent fill
        no_fill = PatternFill(fill_type=None)
        for sheet in checked_wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.fill = no_fill

        rowFill = PatternFill(fgColor="FFCCCB", fill_type="solid")
        for error in self.extractionerror_set.all():
            sheet = checked_wb[error.configuration.worksheet_configuration.sheet_name]
            cell = error.get_cell()
            print("HIGHLIGHT", sheet, cell)
            comment_text =  f"{error.field_name}: {error.error}"
            comment = Comment(comment_text, "Error")
            sheet[cell].fill = rowFill
            sheet[cell].comment = comment
        # save result workbook
        self.file_with_corrections.delete()
        file_name = "%s_%s_checked.xlsx" % (self.pk, self.configuration.code)
        self.file_with_corrections.save(file_name, ContentFile(save_virtual_workbook(checked_wb)))

    def get_organized_extracted_data(self):
        sheet_list = []
        config = self.configuration
        sheets = config.get_sheet_configurations()
        for sheet in sheets:
            section_list = []
            sections = sheet.get_extraction_configurations()
            for section in sections:
                section_data = {
                    "info": section,
                    "extracted_data": section.extracteddata_set.filter(workbook=self).order_by('entry_number')
                }
                section_list.append(section_data)
            sheet_data = {
                "info": sheet,
                "sections": section_list
            }
            sheet_list.append(sheet_data)
        return sheet_list

class ExtractedData(TimeStampModel):
    workbook = models.ForeignKey(Workbook, on_delete=models.CASCADE)
    configuration = models.ForeignKey(DataExtractionConfiguration, on_delete=models.CASCADE)
    data = models.JSONField(encoder=DjangoJSONEncoder)
    entry_number = models.IntegerField(default=0)

    class Meta:
        verbose_name = "Extracted Data"
        verbose_name_plural = "Extracted Data"

    def __str__(self):
        return str(self.workbook.pk) + str(self.configuration.pk) + str(self.entry_number)

    def fields(self):
        return self.data.items()


class ExtractionError(TimeStampModel):
    workbook = models.ForeignKey(Workbook, on_delete=models.CASCADE)
    configuration = models.ForeignKey(DataExtractionConfiguration, default=None, null=True, on_delete=models.SET_NULL)
    field_error = models.BooleanField(default=True)
    entry_number = models.IntegerField(null=True, default=None)
    field_name = models.CharField(max_length=100)
    error = models.TextField()

    class Meta:
        verbose_name = "Extraction Error"
        verbose_name_plural = "Extraction Errors"

    def get_cell(self):
        config = self.configuration
        field = config.datafield_set.filter(field_name=self.field_name)[0]
        if config.extraction_type == DataExtractionConfiguration.TYPE_FORM:
            cell = field.cell
        elif config.extraction_type == DataExtractionConfiguration.TYPE_TABLE_ROWS:
            row = config.min_row + self.entry_number - 1
            column = get_column_letter(field.field_index + config.min_col) #xlsx is 1-based-indexing
            cell = column+str(row)
        print("ERROR CELL", cell)
        return cell