from openpyxl.utils.exceptions import InvalidFileException
from abc import ABCMeta, abstractmethod
from etl.models import Workbook, DataExtractionConfiguration, DataField, ExtractedData, ExtractionError


class ExtractionException(Exception):
    pass


class ExtractionProcess:
    # Generalized Process for Validating and Extracting Data from a worksheet
    __metaclass__ = ABCMeta

    def __init__(self, config, workbook):
        # where workbook is a Workbook model instance
        # where config is a DataExtractionConfiguration instance
        self.workbook = workbook
        self.config = config
        print('CONFIG', self.config)

    def validate_config(self):
        #override me
        return

    def get_worksheet(self):
        # Initialize Worksheet
        wb = self.workbook.get_workbook()
        sheet_name = self.config.sheet_name
        try:
            ws = wb[sheet_name]
        except KeyError:
            message = 'Sheet "' + sheet_name + '" does not exist.'
            raise Workbook.DoesNotExist(message)

        #merged cells cause problems so unmerge them all.
        #for merge in list(ws.merged_cells):
        #    ws.unmerge_cells(range_string=str(merge))
        return ws

    @abstractmethod
    def extract(self):
        ''' To override '''
        #ws = self.get_worksheet()
        #return data
        pass

    def validate(self, data):
        # validate extracted data with given form_class
        # return errors or cleaned_data

        field_rules = self.config.get_field_rules()
        form_class = self.config.get_validation_form()

        form = form_class(data, rules=field_rules, file=self.workbook)

        errors = []
        if form.is_valid():
            return {
                'cleaned_data': form.cleaned_data,
            }
        return {
            'errors': form.errors.as_data()
        }

    def transform(self, data):
        # override me for custom tables
        return data

    def save(self, cleaned_data, commit=True):        

        # transform data before saving
        data = self.transform(cleaned_data)

        # override here for custom tables
        print("SAVE", data)
        extracted_data = ExtractedData(
            workbook = self.workbook,
            configuration = self.config,
            data = data
        )

        if commit: # save data into database
            extracted_data.save()

        return extracted_data

    def save_errors(self, errors, entry_number=None):
        print("ERRORS", errors)
        for field in errors.keys():
            print(field, errors[field])
            for field_error in errors[field]:
                ExtractionError.objects.create(
                    workbook = self.workbook,
                    configuration = self.config,
                    entry_number = entry_number,
                    field_name = field,
                    error = "; ".join(field_error.messages)
                )

    def run(self):
        #extract raw data based on config
        raw_data = self.extract()

        #validate raw data
        data = self.validate(raw_data)
        if 'errors' not in data:
            #save results found into database
            self.save(data['cleaned_data'])
        else:
            errors = data['errors']
            self.save_errors(errors)
            #raise ExtractionException("Errors Found")


class FormDataExtraction(ExtractionProcess):
    #Extract Values from a worksheet formatted as a form
    #config will have a cell field for every field_name

    def extract(self):
        ws = self.get_worksheet()
        fields = self.config.get_field_rules()

        data = {}
        for field in fields:
            value = ws[field.cell].value
            data[field.field_name] = value

        print("RAW", data)
        return data


class RowDataExtraction(ExtractionProcess):
    #Extract values from a worksheet with a table format
    #config will have a field index for each field_name

    def convert_entry_to_values(self, entry):
        entry_values = []
        for cell in entry:
            if cell.hyperlink: # use hyperlink value instead of text value if available
                value = cell.hyperlink.target
            else:
                value = cell.value
            # cast integers as integers. openpyxl reads all numbers as floats.
            if type(value) == float:
                if value.is_integer():
                    value = int(value)
            entry_values.append(value)
        return entry_values

    def convert_entry_to_dict(self, entry):
        fields = self.config.get_field_rules()
        print("FIELDS", fields)
        entry_values = self.convert_entry_to_values(entry)
        print("ROW", entry_values)

        data = {}
        for field in fields:
            print("FIELD", field.field_name, field.field_index)
            data[field.field_name] = entry_values[field.field_index]

        print("CONVERT", data)
        return data

    def extract(self):
        ws = self.get_worksheet()
        min_row = self.config.min_row
        max_row = self.config.max_row
        min_col = self.config.min_col
        max_col = self.config.max_col

        #TO DO VALIDATE RULES WITH NUMBER OF COLUMNS AND COLUMN NAMES
        #BEFORE EXTRACTING

        table_data = []
        entry_number = 0
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            entry_number = entry_number + 1
            if any(cell.value for cell in row): #process only non-blank rows
                data = self.convert_entry_to_dict(row)
                data['_entry_number'] = entry_number
                table_data.append(data)

        return table_data

    def save(self, cleaned_data, commit=True, entry_number=0):
        extracted_data = super().save(cleaned_data, commit=False)
        extracted_data.entry_number = entry_number
        if commit:
            extracted_data.save()
        return extracted_data

    def run(self):
        #extract raw data based on config
        raw_table_data = self.extract()

        for entry in raw_table_data: #raw data is a list, validate per element
            data = self.validate(entry)
            if "cleaned_data" in data:
                cleaned_data = self.transform(data['cleaned_data'])
                self.save(cleaned_data, entry_number=entry['_entry_number'])
            else:
                self.save_errors(data['errors'], entry_number=entry['_entry_number'])


class ColumnDataExtraction(RowDataExtraction):
    #Extract values from a worksheet with a table format where each entry in a column
    #config will have a field_index for each field_name

    def extract(self):
        ws = self.get_worksheet()
        min_row = self.config.min_row
        max_row = self.config.max_row
        min_col = self.config.min_col
        max_col = self.config.max_col

        #TO DO VALIDATE RULES WITH NUMBER OF COLUMNS AND COLUMN NAMES
        #BEFORE EXTRACTING

        table_data = []
        for col in ws.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            if any(cell.value for cell in col): #process only non-blank columns
                data = self.convert_entry_to_dict(col)
                table_data.append(data)

        return table_data
